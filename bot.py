import os
import json
import logging
import re
from datetime import datetime
from pathlib import Path

import httpx
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters, ConversationHandler
)
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN    = os.environ.get("BOT_TOKEN", "")
ADMIN_IDS    = [int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x.strip()]
SALARY_PCT   = float(os.environ.get("SALARY_PCT", "20"))
RATE_MARKUP  = float(os.environ.get("RATE_MARKUP", "2"))
SHEET_ID     = os.environ.get("SHEET_ID", "")
GOOGLE_CREDS = os.environ.get("GOOGLE_CREDS", "")

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

with open("lookup.json", encoding="utf-8") as f:
    LOOKUP: dict = json.load(f)

# Price list (prais EUR) for stock items
try:
    with open("price_lookup.json", encoding="utf-8") as f:
        PRICE_LOOKUP: dict = json.load(f)
except:
    PRICE_LOOKUP: dict = {}

# Invoice-specific lookup: "invoice_num:article" -> record
try:
    with open("invoice_lookup.json", encoding="utf-8") as f:
        INVOICE_LOOKUP: dict = json.load(f)
except:
    INVOICE_LOOKUP: dict = {}

# Storage-only lookup: article -> latest stock record with price
try:
    with open("stock_lookup.json", encoding="utf-8") as f:
        STOCK_LOOKUP: dict = json.load(f)
except:
    STOCK_LOOKUP: dict = {}

# Currency rates archive: "DD.MM.YYYY" -> buy rate
try:
    with open("rates.json", encoding="utf-8") as f:
        RATES: dict = json.load(f)
except:
    RATES: dict = {}

# Client payments per invoice (supports partial payments / tranches):
#   "invoice_num" -> [ {"date": "DD.MM.YYYY", "amount": float, "rate": float}, ... ]
try:
    with open("payments.json", encoding="utf-8") as f:
        PAYMENTS: dict = json.load(f)
except:
    PAYMENTS: dict = {}

# Backward-compat: migrate old payment_dates.json ({inv: "date"}) into tranche form
try:
    with open("payment_dates.json", encoding="utf-8") as f:
        _old_pd = json.load(f)
    for _inv, _d in _old_pd.items():
        _inv = str(_inv).strip()
        if _inv not in PAYMENTS and isinstance(_d, str) and _d:
            PAYMENTS[_inv] = [{"date": _d, "amount": 0.0, "rate": 0.0}]
except:
    pass

_payments_loaded_from_sheet = False  # lazy-load guard (durable store = Google tab)

# Invoice-level meta from orders table: "invoice_num" -> {manager, pay_date, supplier}
try:
    with open("invoice_meta.json", encoding="utf-8") as f:
        INVOICE_META: dict = json.load(f)
except:
    INVOICE_META: dict = {}

# Invoice totals (виторг) captured at PDF-processing time: "65" -> revenue_uah.
# This is the SOURCE OF TRUTH for «Сума рахунку», so we never have to guess it
# from the «ВСІ» sheet (which failed for older invoices).
try:
    with open("invoice_totals.json", encoding="utf-8") as f:
        INVOICE_TOTALS: dict = json.load(f)
except:
    INVOICE_TOTALS: dict = {}

def _save_invoice_totals():
    try:
        with open("invoice_totals.json", "w", encoding="utf-8") as f:
            json.dump(INVOICE_TOTALS, f, ensure_ascii=False)
    except Exception as e:
        logger.error(f"save invoice_totals error: {e}")

def _inv_num(invoice_num_str: str) -> str:
    """Extract the bare invoice number ('65') from 'Рахунок №65 від ...'."""
    m = re.search(r"[№#No]+\s*(\d+)", str(invoice_num_str))
    return m.group(1) if m else ""

def get_rate_from_archive(date_str: str) -> float:
    """Get EUR buy rate from local archive, return 0 if not found"""
    return float(RATES.get(date_str, 0))

def get_payments(inv_num: str) -> list:
    """All payment tranches for an invoice number, or [] if none."""
    if not inv_num:
        return []
    return PAYMENTS.get(str(inv_num).strip(), []) or []

def get_payment_date(inv_num: str) -> str:
    """EARLIEST tranche date for an invoice number, or '' if none."""
    tr = get_payments(inv_num)
    dates = [t.get("date", "") for t in tr if t.get("date")]
    return _earliest_date(*dates) if dates else ""

def total_paid(inv_num: str) -> float:
    return round(sum(float(t.get("amount", 0) or 0) for t in get_payments(inv_num)), 2)

def weighted_rate(inv_num: str) -> float:
    """Payment-amount-weighted rate across tranches; 0 if not computable."""
    tr = [t for t in get_payments(inv_num)
          if float(t.get("amount", 0) or 0) > 0 and float(t.get("rate", 0) or 0) > 0]
    if not tr:
        return 0.0
    num = sum(float(t["amount"]) * float(t["rate"]) for t in tr)
    den = sum(float(t["amount"]) for t in tr)
    return round(num / den, 4) if den else 0.0

def first_payment_rate(inv_num: str) -> float:
    """Rate of the earliest-dated tranche (the base rate the invoice was priced at)."""
    tr = [t for t in get_payments(inv_num) if t.get("date")]
    if not tr:
        return 0.0
    tr = sorted(tr, key=lambda x: _dt(x.get("date", "")))
    for t in tr:
        if float(t.get("rate", 0) or 0) > 0:
            return round(float(t["rate"]), 4)
    return 0.0

def _earliest_date(*dates) -> str:
    """Return earliest of given DD.MM.YYYY dates (ignores blanks)."""
    valid = [d for d in dates if d]
    if not valid:
        return ""
    try:
        return min(valid, key=lambda x: datetime.strptime(x, "%d.%m.%Y"))
    except:
        return valid[0]

def _parse_payment_text(text: str) -> list:
    """Parse Teams-style payment text into tranches.

    Header line:  Оплати за ДД/ММ/РР «...»
    Item lines:   1. КЛІЄНТ рах.834 – 64 531,20
    Header date binds to every рах.НОМЕР below it until the next header.
    Returns list of {"inv": "834", "date": "DD.MM.YYYY", "amount": float}.
    """
    out: list = []
    cur_date = ""
    for line in text.splitlines():
        h = re.search(
            r"[Оо]плат[іиыа]?\s+за\s+(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})",
            line
        )
        if h:
            dd, mm, yy = h.group(1), h.group(2), h.group(3)
            if len(yy) == 2:
                yy = "20" + yy
            try:
                cur_date = f"{int(dd):02d}.{int(mm):02d}.{yy}"
            except:
                cur_date = ""
            continue
        if cur_date:
            # рах.НОМЕР  optionally followed by  – СУМА
            for m in re.finditer(
                r"рах\.?\s*№?\s*(\d+)\s*(?:[–—\-:]\s*([\d][\d\s.,]*))?",
                line, re.IGNORECASE
            ):
                inv = m.group(1)
                amount = 0.0
                if m.group(2):
                    try:
                        s = m.group(2).replace(" ", "").replace("\xa0", "")
                        if s.count(",") and s.count("."):
                            s = s.replace(".", "").replace(",", ".")
                        else:
                            s = s.replace(",", ".")
                        amount = float(s)
                    except:
                        amount = 0.0
                out.append({"inv": inv, "date": cur_date, "amount": amount})
    return out

def _save_payments():
    try:
        with open("payments.json", "w", encoding="utf-8") as f:
            json.dump(PAYMENTS, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"payments save error: {e}")

def _merge_payment_tranches(tranches: list) -> int:
    """Merge tranches (each {inv,date,amount,rate}) into PAYMENTS.
    Dedupe identical (date, amount) for the same invoice. Returns count added."""
    added = 0
    for t in tranches:
        inv = str(t["inv"]).strip()
        lst = PAYMENTS.setdefault(inv, [])
        new_rate = round(float(t.get("rate", 0) or 0), 4)
        existing = None
        for x in lst:
            if x.get("date") == t["date"] and round(float(x.get("amount", 0) or 0), 2) == round(float(t.get("amount", 0) or 0), 2):
                existing = x
                break
        if existing is not None:
            # Same date+amount already stored: refresh the rate if a fresh valid
            # one is available (fixes old tranches frozen at a wrong 52,72 курс).
            if new_rate > 0 and round(float(existing.get("rate", 0) or 0), 4) != new_rate:
                existing["rate"] = new_rate
            continue
        lst.append({"date": t["date"], "amount": round(float(t.get("amount", 0) or 0), 2),
                    "rate": new_rate})
        added += 1
    # keep tranches sorted by date
    for inv, lst in PAYMENTS.items():
        try:
            lst.sort(key=lambda x: datetime.strptime(x.get("date", "01.01.1900"), "%d.%m.%Y"))
        except:
            pass
    _save_payments()
    return added

def get_price(art: str) -> float:
    return PRICE_LOOKUP.get(art, 0)

def shorten_company(name: str) -> str:
    """Convert full legal name to short abbreviation + name"""
    if not name:
        return name
    name = name.strip()
    
    replacements = [
        # Full forms -> abbreviations
        ('ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ', 'ТОВ'),
        ('Товариство з обмеженою відповідальністю', 'ТОВ'),
        ('АКЦІОНЕРНЕ ТОВАРИСТВО', 'АТ'),
        ('Акціонерне товариство', 'АТ'),
        ('ПУБЛІЧНЕ АКЦІОНЕРНЕ ТОВАРИСТВО', 'ПАТ'),
        ('ПРИВАТНЕ АКЦІОНЕРНЕ ТОВАРИСТВО', 'ПРАТ'),
        ('ФІЗИЧНА ОСОБА ПІДПРИЄМЕЦЬ', 'ФОП'),
        ('Фізична особа підприємець', 'ФОП'),
        ('ФІЗИЧНА ОСОБА-ПІДПРИЄМЕЦЬ', 'ФОП'),
        ('Фізична особа-підприємець', 'ФОП'),
        ('ДЕРЖАВНЕ ПІДПРИЄМСТВО', 'ДП'),
        ('КОМУНАЛЬНЕ ПІДПРИЄМСТВО', 'КП'),
        ('ПРИВАТНЕ ПІДПРИЄМСТВО', 'ПП'),
        ('Приватне підприємство', 'ПП'),
    ]
    
    for full, short in replacements:
        if full in name:
            name = name.replace(full, short).strip()
            break
    
    # Remove quotes around company name
    import re as re_co
    name = re_co.sub(r'[\u0022\u201c\u201d\u00ab\u00bb\u2018\u2019]', '', name).strip()
    # Remove extra spaces
    name = ' '.join(name.split())
    
    return name

ORDERS_SHEET_ID = os.environ.get("ORDERS_SHEET_ID", "")
_live_cache: dict = {}
_cache_time = None

def refresh_live_lookup():
    global _live_cache, _cache_time
    # Skip entirely if not configured - we use static lookup.json instead
    if not ORDERS_SHEET_ID:
        return
    import time
    now = time.time()
    # Refresh every 30 minutes
    if _cache_time and now - _cache_time < 1800:
        return
    try:
        creds_dict = json.loads(GOOGLE_CREDS)
        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(ORDERS_SHEET_ID)
        result = {}
        for ws in sh.worksheets():
            src = ws.title
            rows = ws.get_all_values()
            if not rows: continue
            headers = [h.strip() for h in rows[0]]
            # Find column indices
            def col(name):
                for i,h in enumerate(headers):
                    if name.lower() in h.lower(): return i
                return -1
            art_col = col("артикул")
            uktved_col = col("код товару")
            duty_col = col("мито")
            price_col = col("ціна за одиницю")
            weight_col = col("нетто за 1")
            brand_col = col("виробник")
            supplier_col = col("постачальник")
            if art_col < 0: continue
            for row in rows[1:]:
                if len(row) <= art_col: continue
                art = str(row[art_col]).strip()
                if not art: continue
                try: price = float(str(row[price_col]).replace(",",".").replace(" ","")) if price_col>=0 and price_col<len(row) else 0
                except: price = 0
                try: duty = float(str(row[duty_col]).replace(",",".").replace(" ","").replace("%","")) if duty_col>=0 and duty_col<len(row) else 0.04
                except: duty = 0.04
                if duty > 1: duty = duty / 100
                try: weight = float(str(row[weight_col]).replace(",",".").replace(" ","")) if weight_col>=0 and weight_col<len(row) else 0
                except: weight = 0
                result[art] = {
                    "uktved": str(row[uktved_col]).strip() if uktved_col>=0 and uktved_col<len(row) else "",
                    "duty": duty,
                    "cost_eur": price,
                    "weight": weight,
                    "brand": str(row[brand_col]).strip() if brand_col>=0 and brand_col<len(row) else "",
                    "supplier": str(row[supplier_col]).strip() if supplier_col>=0 and supplier_col<len(row) else "",
                    "source": src,
                }
        _live_cache = result
        _cache_time = now
        logger.info(f"Live lookup refreshed: {len(result)} articles")
    except Exception as e:
        logger.error(f"Live lookup error: {e}")

# ── Google Sheets ─────────────────────────────────────────────────────────────
# Cache gspread client to avoid re-auth on every call
_gsheet_cache = None
_gsheet_time = 0

def get_gsheet():
    global _gsheet_cache, _gsheet_time
    import time
    now = time.time()
    # Cache for 30 min - tokens last 1 hour
    if _gsheet_cache and now - _gsheet_time < 1800:
        return _gsheet_cache
    try:
        creds_dict = json.loads(GOOGLE_CREDS)
        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        gc = gspread.authorize(creds)
        _gsheet_cache = gc.open_by_key(SHEET_ID)
        _gsheet_time = now
        return _gsheet_cache
    except Exception as e:
        logger.error(f"GSheet connect error: {e}")
        _gsheet_cache = None
        return None

def get_or_create_ws(spreadsheet, name: str):
    try:
        return spreadsheet.worksheet(name)
    except gspread.WorksheetNotFound:
        pass
    try:
        ws = spreadsheet.add_worksheet(title=name, rows=500, cols=20)
    except gspread.exceptions.APIError as e:
        if "already exists" in str(e):
            return spreadsheet.worksheet(name)
        raise
    headers = ["Менеджер","Клієнт","Рахунок","Дата оплати","Артикул",
               "Кть","Закуп EUR","Мито%","Курс",
               "Собів UAH/шт","Собів загал",
               "Ціна прод UAH","Виторг",
               "Прибуток (S)","Надбавка (T)","Склад?",
               "УКТЗЕД","Бренд","Джерело",
               "Прайс EUR","Вага/шт","Вага Китай","Вага Європа","Додано","Постачальник"]
    ws.append_row(headers)
    ws.format("A1:Y1", {
        "backgroundColor": {"red":0.17,"green":0.18,"blue":0.24},
        "textFormat": {"bold":True,"foregroundColor":{"red":1,"green":1,"blue":1}},
        "horizontalAlignment": "CENTER"
    })
    return ws

def append_to_sheets(manager_name: str, inv: dict):
    """Write invoice to Google Sheets using BATCH updates for speed."""
    try:
        sh = get_gsheet()
        if not sh:
            return False
        ws_mgr = get_or_create_ws(sh, manager_name)
        ws_all = get_or_create_ws(sh, "ВСІ")

        rate = inv.get("rate", 0)
        client = inv.get("client", "")
        invoice_num = inv.get("invoice_num", "")
        date = inv.get("date", "")
        items = inv.get("items", [])
        import re as re_sh
        m_sh = re_sh.search(r"[№#No]+\s*(\d+)", invoice_num)
        inv_number = m_sh.group(1) if m_sh else ""

        # Get current row count
        existing = ws_mgr.get_all_values()
        start_row = len(existing) + 1

        # Build ALL rows in memory first
        rows_to_write = []
        format_requests = []
        all_sheet_rows = []  # for ВСІ

        # Separator row (if not first entry)
        if start_row > 2:
            rows_to_write.append([""] * 25)
            format_requests.append({
                "range": f"A{start_row}:Y{start_row}",
                "format": {"backgroundColor": {"red": 0.85, "green": 0.91, "blue": 0.97}}
            })
            all_sheet_rows.append([""] * 25)
            start_row += 1

        # Invoice header row
        header = [manager_name, client, invoice_num, date] + [""] * 21
        rows_to_write.append(header)
        format_requests.append({
            "range": f"A{start_row}:Y{start_row}",
            "format": {
                "backgroundColor": {"red": 0.78, "green": 0.87, "blue": 0.95},
                "textFormat": {"bold": True}
            }
        })
        all_sheet_rows.append(header)
        start_row += 1

        # Item rows with formulas
        stock_rows = []  # rows to color red on E column
        for item in items:
            lu = lookup_article(item["article"], inv_number, item.get("is_stock", False))
            cost_eur = lu.get("cost_eur", 0)
            duty = lu.get("duty", 0.04)
            price_eur = get_price(item["article"])
            weight_unit = lu.get("weight", 0)
            source = lu.get("source", "")
            is_stock = item.get("is_stock", False)
            duty_pct = round(duty * 100, 1)
            r = start_row

            if is_stock:
                fn = f"=M{r}-T{r}*I{r}*F{r}"
                fo = f"=(M{r}-K{r})-N{r}"
            else:
                fn = f"=M{r}-K{r}"
                fo = "=0"

            row = [
                "",                                              # A Менеджер
                "",                                              # B Клієнт
                "",                                              # C Рахунок
                "",                                              # D Дата
                item["article"],                                 # E Артикул
                item["qty"],                                     # F Кть
                round(cost_eur, 2),                              # G Закуп EUR
                duty_pct,                                        # H Мито%
                round(rate, 2) if rate else "",                  # I Курс
                f"=G{r}*(1+H{r}/100)*I{r}",                      # J Собів UAH/шт
                f"=J{r}*F{r}",                                   # K Собів загал
                item["price_uah"],                               # L Ціна прод
                f"=L{r}*F{r}",                                   # M Виторг
                fn,                                              # N Прибуток S
                fo,                                              # O Надбавка T
                "так" if is_stock else "",                       # P Склад?
                lu.get("uktved", ""),                            # Q УКТЗЕД
                lu.get("brand", ""),                             # R Бренд
                source,                                          # S Джерело
                round(price_eur, 2) if price_eur else "",        # T Прайс EUR
                round(weight_unit, 3) if weight_unit else "",    # U Вага/шт
                round(weight_unit * item["qty"], 3) if (is_stock and source == "Китай") else 0,  # V Вага Китай
                round(weight_unit * item["qty"], 3) if (is_stock and "trade" in source.lower()) else 0,  # W Вага Європа
                datetime.now().strftime("%d.%m.%Y %H:%M"),       # X Додано
                lu.get("supplier", ""),                          # Y Постачальник
            ]
            rows_to_write.append(row)
            all_sheet_rows.append(row)

            if is_stock:
                stock_rows.append(r)

            start_row += 1

        # ── ONE batch write to manager sheet ──────────────────────────────────
        first_row = len(existing) + 1
        last_row = first_row + len(rows_to_write) - 1
        ws_mgr.update(
            f"A{first_row}:Y{last_row}",
            rows_to_write,
            value_input_option="USER_ENTERED"
        )

        # ── ALL formatting in ONE batch_update call (massive speedup) ─────
        batch_formats = []

        # Number formatting for data rows
        num_format = {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.00"}}
        data_start = first_row + (2 if existing and len(existing) > 1 else 1)
        if data_start <= last_row:
            for col in ["G", "H", "I", "J", "K", "L", "M", "N", "O", "T", "U", "V", "W"]:
                batch_formats.append({"range": f"{col}{data_start}:{col}{last_row}", "format": num_format})

        # Stock article cells - red
        stock_fmt = {
            "backgroundColor": {"red": 0.99, "green": 0.87, "blue": 0.87},
            "textFormat": {"bold": True, "foregroundColor": {"red": 0.8, "green": 0.0, "blue": 0.0}}
        }
        for sr in stock_rows:
            batch_formats.append({"range": f"E{sr}", "format": stock_fmt})

        # Separator/header rows
        batch_formats.extend(format_requests)

        # Apply ALL formats in ONE API call
        if batch_formats:
            try:
                ws_mgr.batch_format(batch_formats)
            except Exception as e:
                # Fallback to individual format calls if batch_format unsupported
                logger.warning(f"batch_format failed, fallback: {e}")
                for bf in batch_formats:
                    try:
                        ws_mgr.format(bf["range"], bf["format"])
                    except: pass

        # ── Append to ВСІ sheet (single batch) ────────────────────────────────
        try:
            ws_all.append_rows(all_sheet_rows, value_input_option="USER_ENTERED")
        except Exception as e:
            logger.warning(f"ВСІ append error: {e}")

        return True
    except Exception as e:
        logger.error(f"Sheet append error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

# ── Storage ───────────────────────────────────────────────────────────────────
def _uf(uid): return DATA_DIR / f"{uid}.json"
def load_user(uid):
    p = _uf(uid)
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else {"name":"","invoices":[]}
def save_user(uid, data):
    _uf(uid).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Currency rate (archive first, then Minfin, then NBU) ─────────────────────
async def get_nbu_rate(date_str: str):
    """Get EUR buy rate: archive -> Minfin -> NBU fallback"""
    import re as re_rate

    # ── Method 0: Local rates archive (most accurate) ─────────────────────────
    archived = get_rate_from_archive(date_str)
    if archived > 0:
        final = round(archived * (1 + RATE_MARKUP / 100), 2)
        logger.info(f"Archive rate {date_str}: {archived} -> {final}")
        return final

    try:
        dt = datetime.strptime(date_str, "%d.%m.%Y")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "uk-UA,uk;q=0.9,ru;q=0.8",
            "Referer": "https://minfin.com.ua/currency/mb/",
        }

        # ── Method 1: Minfin EUR archive page ────────────────────────────────
        # URL: /currency/mb/eur/DD-MM-YYYY/
        page_url = f"https://minfin.com.ua/currency/mb/eur/{dt.day:02d}-{dt.month:02d}-{dt.year}/"
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            r = await client.get(page_url, headers=headers)
            text = r.text

        # Look for: "по курсу межбанка 51,4287"
        m = re_rate.search(r"по курсу межбанка\s+([\d]+[,.][\d]+)", text)
        if m:
            rate = float(m.group(1).replace(",", "."))
            if 48 < rate < 65:  # EUR/UAH realistic range
                final = round(rate * (1 + RATE_MARKUP / 100), 2)
                logger.info(f"Minfin EUR buy {date_str}: {rate} -> {final}")
                return final

        # Look for JSON in __NEXT_DATA__ or similar script tags
        # Pattern: "buy":"51.4287" near "eur" context
        eur_section = text.lower()
        eur_pos = eur_section.find('"eur"')
        if eur_pos == -1:
            eur_pos = eur_section.find("euro")
        if eur_pos > 0:
            chunk = text[max(0, eur_pos-200):eur_pos+500]
            m = re_rate.search(r'"buy"\s*:\s*"?([\d.]+)"?', chunk)
            if m:
                rate = float(m.group(1))
                if 48 < rate < 65:
                    final = round(rate * (1 + RATE_MARKUP / 100), 2)
                    logger.info(f"Minfin JSON EUR buy {date_str}: {rate} -> {final}")
                    return final

        # ── Method 2: search all "buy" values, pick EUR range ────────────────
        all_buys = re_rate.findall(r'"buy"\s*:\s*"?([\d.]+)"?', text)
        for b in all_buys:
            rate = float(b)
            if 48 < rate < 65:  # EUR range
                final = round(rate * (1 + RATE_MARKUP / 100), 2)
                logger.info(f"Minfin all-buy EUR {date_str}: {rate} -> {final}")
                return final

        # ── Method 3: find any number in EUR range on EUR page ───────────────
        all_nums = re_rate.findall(r"5[0-9][,.][\d]{4}", text)
        if all_nums:
            rate = float(all_nums[0].replace(",", "."))
            if 48 < rate < 65:
                final = round(rate * (1 + RATE_MARKUP / 100), 2)
                logger.info(f"Minfin num EUR {date_str}: {rate} -> {final}")
                return final

    except Exception as e:
        logger.warning(f"Minfin rate error: {e}")

    # ── Fallback: NBU official rate ───────────────────────────────────────────
    try:
        dt = datetime.strptime(date_str, "%d.%m.%Y")
        url = f"https://bank.gov.ua/NBU_Exchange/exchange_site?start={dt:%Y%m%d}&end={dt:%Y%m%d}&valcode=EUR&sort=exchangedate&order=desc&json"
        async with httpx.AsyncClient(timeout=8) as client:
            r = await client.get(url)
            data = r.json()
            if data:
                rate = float(data[0]["rate"])
                logger.info(f"NBU fallback EUR {date_str}: {rate}")
                return round(rate * (1 + RATE_MARKUP / 100), 2)
    except Exception as e:
        logger.warning(f"NBU fallback error: {e}")

    return None


async def get_payment_rate(date_str: str):
    """Rate for a PAYMENT date — used for 50/50 коригування.

    archive (межбанк, +markup)  →  NBU official by EXACT date (+markup).

    Deliberately does NOT use the greedy Minfin scrape: for dates outside the
    local archive that scrape grabs the *current* rate (e.g. 52,72 for every
    old date), which made both tranches identical and zeroed out коригування.
    NBU is date-accurate, so distinct payment dates get distinct rates.
    """
    a = get_rate_from_archive(date_str)
    if a > 0:
        return round(a * (1 + RATE_MARKUP / 100), 2)
    try:
        dt = datetime.strptime(date_str, "%d.%m.%Y")
        url = (f"https://bank.gov.ua/NBU_Exchange/exchange_site?"
               f"start={dt:%Y%m%d}&end={dt:%Y%m%d}&valcode=EUR"
               f"&sort=exchangedate&order=desc&json")
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(url)
            data = r.json()
            if data:
                rate = float(data[0]["rate"])
                final = round(rate * (1 + RATE_MARKUP / 100), 2)
                logger.info(f"Payment rate NBU {date_str}: {rate} -> {final}")
                return final
    except Exception as e:
        logger.warning(f"payment rate error {date_str}: {e}")
    return None


# ── PDF parser ────────────────────────────────────────────────────────────────
def parse_pdf(path: str) -> dict:
    result = {"invoice_num":"","date":"","client":"","items":[]}
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        logger.error(f"PDF error: {e}")
        return result

    m = re.search(r"Рахунок на оплату\s*[№#]\s*(\d+)\s*від\s*([\d]+\s+\w+\s+\d{4})", text)
    if m:
        result["invoice_num"] = f"Рахунок №{m.group(1)} від {m.group(2)}"
        months = {"січня":1,"лютого":2,"березня":3,"квітня":4,"травня":5,"червня":6,
                  "липня":7,"серпня":8,"вересня":9,"жовтня":10,"листопада":11,"грудня":12}
        dm = re.search(r"(\d+)\s+(\w+)\s+(\d{4})", m.group(2))
        if dm:
            mn = months.get(dm.group(2).lower(), 0)
            if mn:
                result["date"] = f"{int(dm.group(1)):02d}.{mn:02d}.{dm.group(3)}"

    m2 = re.search(r"Покупець:\s*(.+?)(?:\n|Тел)", text, re.DOTALL)
    if m2:
        result["client"] = shorten_company(m2.group(1).strip()[:120])

    # ── Method 1: Find articles in PDF text PRESERVING ORDER from PDF ──────
    # Scan text positions, longer articles take precedence on overlap
    known_set = set(LOOKUP.keys())
    found_items = []  # list of (text_position, article, qty, price)
    seen_positions = set()

    # Sort articles by length DESCENDING so longer ones match first (avoid substring matches)
    sorted_known = sorted(known_set, key=len, reverse=True)

    for art in sorted_known:
        art_escaped = re.escape(art)
        pattern = art_escaped + r"\s+(\d+)\s*шт\s+([\d\s]+[,.]\d{2})"
        for m in re.finditer(pattern, text):
            pos = m.start()
            # Skip if this position overlaps with already-found article
            if any(abs(pos - p) < len(art) for p in seen_positions):
                continue
            try:
                qty = int(m.group(1))
                price = float(m.group(2).replace(" ","").replace(",","."))
                if qty > 0 and price > 0:
                    found_items.append((pos, art, qty, price))
                    seen_positions.add(pos)
                    break  # one match per article
            except: pass

    # Sort by position in text to preserve PDF order
    found_items.sort(key=lambda x: x[0])
    for pos, art, qty, price in found_items:
        result["items"].append({"article": art, "qty": qty, "price_uah": price})

    # ── Method 2: Generic regex for any article-looking pattern (fallback) ───
    if not result["items"]:
        # Match line format: "№ Назва товару АРТИКУЛ qty шт price total"
        # Articles can contain: A-Z, 0-9, -, /, +, ., (, )
        item_pat = re.compile(
            r"\d{1,3}\s+[\w\s,\-\'\.]+?"
            r"([A-Z0-9][A-Z0-9\-\/\.\+\(\)]{4,})"
            r"\s+(\d+)\s+шт\s+([\d\s]+[,.]\d{2})\s+([\d\s]+[,.]\d{2})"
        )
        for m in item_pat.finditer(text):
            art = m.group(1).strip()
            if art in seen: continue
            seen.add(art)
            try:
                price = float(m.group(3).replace(" ","").replace(",","."))
                result["items"].append({"article":art,"qty":int(m.group(2)),"price_uah":price})
            except: pass

        if not result["items"]:
            for m in re.finditer(r"([A-Z0-9][A-Z0-9\-\/\.\+\(\)]{5,})\s+(\d{1,3})\s+шт\s+([\d\s]+[,.]\d{2})", text):
                try:
                    result["items"].append({"article":m.group(1).strip(),"qty":int(m.group(2)),
                                            "price_uah":float(m.group(3).replace(" ","").replace(",","."))})
                except: pass
    return result

def lookup_article(art: str, invoice_num: str = "", is_stock: bool = False) -> dict:
    """Find article record. INVOICE MATCH ALWAYS WINS over stock_lookup."""
    art = art.strip()

    # 1) Invoice-specific match - ALWAYS highest priority
    # If we have a record for THIS exact invoice, use it (regardless of stock flag)
    if invoice_num:
        inv_key = f"{invoice_num}:{art}"
        if inv_key in INVOICE_LOOKUP:
            rec = INVOICE_LOOKUP[inv_key]
            if rec.get("cost_eur", 0) > 0:
                return rec

    # 2) Stock lookup - only if marked as stock by user
    if is_stock and art in STOCK_LOOKUP:
        rec = STOCK_LOOKUP[art]
        if rec.get("cost_eur", 0) > 0:
            return rec

    # 3) General lookup (latest with price)
    if art in LOOKUP:
        rec = LOOKUP[art]
        if rec.get("cost_eur", 0) > 0:
            return rec

    # 4) Stock lookup fallback (even if not marked)
    if art in STOCK_LOOKUP:
        rec = STOCK_LOOKUP[art]
        if rec.get("cost_eur", 0) > 0:
            return rec

    # 5) Invoice match with 0 price (item in transit)
    if invoice_num:
        inv_key = f"{invoice_num}:{art}"
        if inv_key in INVOICE_LOOKUP:
            return INVOICE_LOOKUP[inv_key]

    # 6) Live cache
    if ORDERS_SHEET_ID and _live_cache:
        result = _live_cache.get(art)
        if result:
            return result

    # 7) Fuzzy match
    art_norm = " ".join(art.upper().split())
    for key, val in LOOKUP.items():
        if " ".join(key.upper().split()) == art_norm:
            return val

    return LOOKUP.get(art, {})

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(manager_name, invoices, month):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month
    thin = Side(style='thin', color='BDC3C7')
    brd = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells('A1:N1')
    ws['A1'] = f'РОЗРАХУНОК ЗП — {manager_name} — {month}'
    ws['A1'].font = Font(name='Arial',bold=True,size=13,color='FFFFFF')
    ws['A1'].fill = PatternFill('solid',fgColor='1F2D3D')
    ws['A1'].alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 26

    headers = ['Клієнт','Рахунок','Дата','Артикул','Кть','Закуп EUR','Мито%',
               'Курс','Собів UAH','Ціна UAH','Виторг','Собів загал',
               'Прибуток S','Надбавка T','Склад?','Прайс EUR','Вага/шт','Вага Китай','Вага Європа']
    ws.row_dimensions[2].height = 36
    for c,h in enumerate(headers,1):
        cell = ws.cell(2,c,h)
        cell.font = Font(name='Arial',bold=True,size=8,color='FFFFFF')
        cell.fill = PatternFill('solid',fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = brd

    for i,w in enumerate([18,22,12,22,5,10,7,10,12,14,13,13,13,8],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for inv in invoices:
        import re as re_b
        m_b = re_b.search(r"[№#No]+\s*(\d+)", inv.get("invoice_num",""))
        inv_num_b = m_b.group(1) if m_b else ""
        for item in inv.get("items",[]):
            lu = lookup_article(item["article"], inv_num_b, item.get("is_stock", False))
            cost_eur = lu.get("cost_eur",0)
            duty = lu.get("duty",0.04)
            rate = item.get("rate", inv.get("rate",52.0))
            cost_unit = cost_eur*(1+duty)*rate
            revenue = item["price_uah"]*item["qty"]
            cost_total = cost_unit*item["qty"]
            profit = (item["price_uah"]-cost_eur*(1+duty)*rate)*item["qty"] if item.get("is_stock") else revenue-cost_total
            is_stock = item.get("is_stock",False)
            price_eur = get_price(item["article"])
            w_unit = lu.get("weight",0)
            source = lu.get("source","")
            wx = round(revenue - price_eur*rate*item["qty"],2) if (is_stock and price_eur) else 0
            s_val = wx if is_stock else round(profit,2)
            t_val = round((revenue-cost_total)-wx,2) if is_stock else 0
            w_china = round(w_unit*item["qty"],3) if (is_stock and "китай" in source.lower()) else 0
            w_eu    = round(w_unit*item["qty"],3) if (is_stock and "e-trade" in source.lower()) else 0
            vals = [inv.get("client",""),inv.get("invoice_num",""),inv.get("date",""),
                    item["article"],item["qty"],cost_eur,f'{duty*100:.0f}%',rate,
                    round(cost_unit,2),item["price_uah"],round(revenue,2),round(cost_total,2),
                    s_val,t_val,"так" if is_stock else "",price_eur,round(w_unit,3),w_china,w_eu]
            for c,v in enumerate(vals,1):
                cell = ws.cell(row,c,v)
                cell.font = Font(name='Arial',size=9)
                cell.border = brd
                cell.alignment = Alignment(vertical='center')
                if c in [6,8,9,10,11,12,13]: cell.number_format = '#,##0.00'
                cell.fill = PatternFill('solid',fgColor='FDECEA' if is_stock else ('F8F9FA' if row%2==0 else 'FFFFFF'))
            ws.row_dimensions[row].height = 17
            row += 1

    TOT = row
    ws.merge_cells(f'A{TOT}:J{TOT}')
    ws.cell(TOT,1,'ПІДСУМОК').font = Font(name='Arial',bold=True,size=10,color='FFFFFF')
    ws.cell(TOT,1).fill = PatternFill('solid',fgColor='1A5276')
    ws.cell(TOT,1).alignment = Alignment(horizontal='right')
    for c in [11,12,13]:
        cl = get_column_letter(c)
        cell = ws.cell(TOT,c,f'=SUM({cl}3:{cl}{TOT-1})')
        cell.font = Font(name='Arial',bold=True,size=10,color='FFFFFF')
        cell.fill = PatternFill('solid',fgColor='1A5276')
        cell.number_format = '#,##0.00'
        cell.border = brd

    SR = TOT+2
    rc = get_column_letter(13)
    labels = [('Валовий прибуток (грн):',f'={rc}{TOT}'),
              ('▶ Витрати на доставку (вручну):',0),
              ('Чистий прибуток (грн):',f'=D{SR}-D{SR+1}'),
              (f'ЗП ({SALARY_PCT}% від чистого прибутку):',f'=MAX(0,D{SR+2})*{SALARY_PCT}/100')]
    for i,(lbl,val) in enumerate(labels):
        r2=SR+i
        ws.merge_cells(f'A{r2}:C{r2}')
        ws.cell(r2,1,lbl).font=Font(name='Arial',size=10,bold=(i in[1,2,3]))
        ws.cell(r2,1).alignment=Alignment(horizontal='right')
        cv=ws.cell(r2,4,val)
        cv.number_format='#,##0.00'
        if i==1:
            cv.fill=PatternFill('solid',fgColor='EBF5FB')
            cv.font=Font(name='Arial',size=11,color='000080',bold=True)
        elif i==3:
            cv.fill=PatternFill('solid',fgColor='D5F5E3')
            cv.font=Font(name='Arial',size=14,color='1E8449',bold=True)

    path = str(DATA_DIR/f"salary_{manager_name}_{month}.xlsx")
    wb.save(path)
    return path

# ── States ────────────────────────────────────────────────────────────────────
WAIT_NAME, WAIT_DATE, WAIT_STOCK, WAIT_DELIVERY, WAIT_EXCEL, WAIT_RATES, WAIT_PAYMENTS, WAIT_SUMA = range(8)

# ── Handlers ──────────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user = load_user(update.effective_user.id)
    name = user.get("name","")
    if name:
        await update.message.reply_text(
            f"👋 Привіт, {name}!\n\n"
            "Надішли PDF рахунку — я його оброблю.\n\n"
            "Команди:\n"
            "/report — Excel з ЗП за місяць\n"
            "/clear — очистити місяць\n"
            "/name — змінити ім'я\n"
            "/sheet — посилання на Google таблицю"
            + ("\n/admin — всі менеджери\n/update — оновити довідник\n/rates — оновити курси валют\n/oplata — завантажити дати оплат\n/suma — вписати суму рахунку (старі рах.)\n/perekurs — перерахувати курси оплат" if update.effective_user.id in ADMIN_IDS else "")
        )
    else:
        await update.message.reply_text("👋 Привіт! Як тебе звати? (введи ім'я)")
        return WAIT_NAME

async def set_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    user = load_user(update.effective_user.id)
    user["name"] = name
    save_user(update.effective_user.id, user)
    await update.message.reply_text(f"✅ Збережено! Привіт, {name}!\n\nНадсилай PDF рахунки.")
    return ConversationHandler.END

async def handle_pdf(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user = load_user(uid)
    if not user.get("name"):
        await update.message.reply_text("Спочатку введи ім'я: /start")
        return

    refresh_live_lookup()
    msg = await update.message.reply_text("⏳ Обробляю PDF...")
    doc = update.message.document
    file = await ctx.bot.get_file(doc.file_id)
    pdf_path = str(DATA_DIR/f"{uid}_{doc.file_id}.pdf")
    await file.download_to_drive(pdf_path)

    parsed = parse_pdf(pdf_path)
    os.remove(pdf_path)

    if not parsed["items"]:
        await msg.edit_text("❌ Не вдалось розпізнати позиції в PDF.")
        return

    # Extract invoice number for precise lookup
    inv_number = ""
    import re as re_inv
    m_inv = re_inv.search(r'[№#No]+\s*(\d+)', parsed.get("invoice_num", ""))
    if m_inv:
        inv_number = m_inv.group(1)

    found, not_found = [], []
    for item in parsed["items"]:
        # Try with invoice match AND stock fallback (assume might be stock)
        lu = lookup_article(item["article"], inv_number)
        if not lu.get("cost_eur"):
            # Try as stock item (checks stock_lookup)
            lu = lookup_article(item["article"], inv_number, is_stock=True)
        if lu.get("cost_eur"):
            item.update(lu)
            found.append(item["article"])
        else:
            not_found.append(item["article"])

    ctx.user_data["pending_invoice"] = {
        "invoice_num": parsed["invoice_num"],
        "date": parsed["date"],
        "client": parsed["client"],
        "items": parsed["items"],
    }

    lines = [f"📄 *{parsed['invoice_num']}*", f"👤 {parsed['client']}", ""]
    for item in parsed["items"]:
        # Use same fallback chain as save logic
        lu = lookup_article(item["article"], inv_number)
        if not lu.get("cost_eur"):
            lu = lookup_article(item["article"], inv_number, is_stock=True)
        cost = f"{lu['cost_eur']:.2f} EUR" if lu.get("cost_eur") else "❓ немає в довіднику"
        lines.append(f"• `{item['article']}` × {item['qty']} — {item['price_uah']:,.2f} грн | Закуп: {cost}")

    lines.append(f"\n✅ Знайдено: {len(found)}/{len(parsed['items'])}")
    if not_found:
        lines.append(f"⚠️ Не знайдено ({len(not_found)}): {', '.join(not_found)}")

    # ── Date / rate / manager logic ──────────────────────────────────────────
    # Date sources for THIS invoice:
    #   table_pay     — "Дата оплати клієнта" (AA) from orders table  ← PRIORITY
    #   oplata_pay    — actual payment date loaded via /oplata (double-check)
    #   table_confirm — "Дата підтвердження замовлення" (V), fallback only
    # Rule: rate date = EARLIEST of (table_pay, oplata_pay); else order-confirm date.
    table_pay = ""
    table_confirm = ""
    auto_manager = ""
    if inv_number:
        meta = INVOICE_META.get(inv_number, {})
        table_pay = meta.get("pay_date", "") or ""
        auto_manager = meta.get("manager", "") or ""
        for item in parsed["items"]:
            rec = INVOICE_LOOKUP.get(f"{inv_number}:{item['article']}")
            if rec:
                if not table_pay:     table_pay = rec.get("pay_date", "") or ""
                if not table_confirm: table_confirm = rec.get("confirm_date", "") or ""
                if not auto_manager:  auto_manager = rec.get("manager", "") or ""

    _ensure_payments_loaded()
    oplata_pay = get_payment_date(inv_number)
    pay_eff = _earliest_date(table_pay, oplata_pay)
    auto_date = pay_eff or table_confirm

    ctx.user_data["pending_invoice"]["auto_manager"] = auto_manager
    if auto_manager:
        lines.append(f"👤 Менеджер (з таблиці): *{auto_manager}*")

    # Note explaining which date was chosen
    if table_pay and oplata_pay and table_pay != oplata_pay:
        date_note = f" _(найраніша з: таблиця {table_pay} / оплата {oplata_pay})_"
    elif oplata_pay and not table_pay:
        date_note = " _(дата оплати з /oplata)_"
    elif table_pay:
        date_note = " _(дата оплати з таблиці)_"
    elif table_confirm:
        date_note = " _(дата підтвердження замовлення)_"
    else:
        date_note = ""

    items = ctx.user_data["pending_invoice"].get("items", [])
    ctx.user_data["stock_selected"] = set()

    if auto_date:
        # Have a date (from table and/or actual payment) - use it automatically
        rate = await get_nbu_rate(auto_date)
        if rate:
            ctx.user_data["pending_invoice"]["rate"] = rate
            ctx.user_data["pending_invoice"]["date"] = auto_date
            for item in ctx.user_data["pending_invoice"]["items"]:
                item["rate"] = rate
            lines.append(f"\n💱 Дата: *{auto_date}*{date_note} | Курс: *{rate:.2f} грн/EUR* (+{RATE_MARKUP}%)")
        else:
            ctx.user_data["pending_invoice"]["rate"] = 0
            ctx.user_data["pending_invoice"]["date"] = auto_date
            lines.append(f"\n💱 Дата: *{auto_date}*{date_note} | ⚠️ Курс не знайдено — вкажи вручну в таблиці")

        await msg.edit_text("\n".join(lines), parse_mode="Markdown")
        await msg.reply_text(
            "📦 *Вибери складські товари* (натисни щоб відмітити):",
            reply_markup=build_stock_keyboard(items, set()),
            parse_mode="Markdown"
        )
        return WAIT_STOCK
    else:
        # No invoice-specific records found - all stock, ask manager for date
        ctx.user_data["pending_invoice"]["rate"] = 0
        ctx.user_data["pending_invoice"]["date"] = ""
        lines.append(f"\n📅 Всі товари зі складу — введи дату оплати клієнтом (ДД.ММ.РРРР):")
        await msg.edit_text("\n".join(lines), parse_mode="Markdown")
        return WAIT_DATE

async def handle_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    date_str = update.message.text.strip()
    if not re.match(r"\d{2}\.\d{2}\.\d{4}", date_str):
        await update.message.reply_text("❌ Формат: ДД.ММ.РРРР (наприклад 23.04.2026)")
        return WAIT_DATE

    ctx.user_data["pending_invoice"]["date"] = date_str
    rate = await get_nbu_rate(date_str)
    if rate:
        ctx.user_data["pending_invoice"]["rate"] = rate
        for item in ctx.user_data["pending_invoice"]["items"]:
            item["rate"] = rate
        rate_msg = f"💱 Курс міжбанк (Мінфін) на {date_str}: *{rate:.2f} грн/EUR* (+{RATE_MARKUP}%)"
    else:
        ctx.user_data["pending_invoice"]["rate"] = 52.0
        rate_msg = f"⚠️ Не вдалось отримати курс. Використовую 52.00"

    inv = ctx.user_data.get("pending_invoice", {})
    items = inv.get("items", [])
    ctx.user_data["stock_selected"] = set()
    await update.message.reply_text(rate_msg, parse_mode="Markdown")
    await update.message.reply_text(
        "📦 *Вибери складські товари* (натисни щоб відмітити):\n_(для складського рахунку можна одразу натиснути Зберегти)_",
        reply_markup=build_stock_keyboard(items, set()),
        parse_mode="Markdown"
    )
    return WAIT_STOCK

def build_stock_keyboard(items: list, selected: set) -> InlineKeyboardMarkup:
    """Build keyboard with checkboxes for each item."""
    buttons = []
    for i, item in enumerate(items):
        art = item["article"]
        qty = item["qty"]
        check = "🔴" if art in selected else "⚪"
        # Use index instead of article name to keep callback_data short
        buttons.append([InlineKeyboardButton(
            f"{check} {art} × {qty}",
            callback_data=f"stk_{i}"
        )])
    # Action buttons
    all_selected = len(selected) == len(items) and len(items) > 0
    buttons.append([
        InlineKeyboardButton("🔴 Все склад" if not all_selected else "⚪ Зняти все",
                             callback_data="stk_all"),
    ])
    buttons.append([
        InlineKeyboardButton("✅ Зберегти", callback_data="stk_save"),
    ])
    return InlineKeyboardMarkup(buttons)

async def handle_stock(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    inv = ctx.user_data.get("pending_invoice", {})
    items = inv.get("items", [])
    ctx.user_data["stock_selected"] = set()

    await update.message.reply_text(
        "📦 *Вибери складські товари* (натисни щоб відмітити):",
        reply_markup=build_stock_keyboard(items, set()),
        parse_mode="Markdown"
    )
    return WAIT_STOCK

async def callback_stock(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    inv = ctx.user_data.get("pending_invoice", {})
    items = inv.get("items", [])
    selected = ctx.user_data.get("stock_selected", set())

    if query.data == "stk_save":
        await save_invoice(query, ctx)
        return ConversationHandler.END

    if query.data == "stk_all":
        # Toggle all
        all_articles = {item["article"] for item in items}
        if selected == all_articles:
            selected = set()
        else:
            selected = all_articles
        ctx.user_data["stock_selected"] = selected
        await query.edit_message_reply_markup(
            reply_markup=build_stock_keyboard(items, selected)
        )
        return WAIT_STOCK

    if query.data.startswith("stk_"):
        try:
            idx = int(query.data[4:])
            if 0 <= idx < len(items):
                art = items[idx]["article"]
                if art in selected:
                    selected.discard(art)
                else:
                    selected.add(art)
                ctx.user_data["stock_selected"] = selected
                await query.edit_message_reply_markup(
                    reply_markup=build_stock_keyboard(items, selected)
                )
        except: pass

async def save_invoice(query, ctx: ContextTypes.DEFAULT_TYPE):
    uid = query.from_user.id
    user = load_user(uid)
    inv = ctx.user_data.get("pending_invoice", {})
    selected = ctx.user_data.get("stock_selected", set())

    # Mark stock items
    for item in inv.get("items", []):
        item["is_stock"] = item["article"] in selected

    user.setdefault("invoices", []).append(inv)
    save_user(uid, user)

    # Manager: prefer the one detected from the orders table (by invoice number),
    # so the invoice lands in the right manager's sheet regardless of who sent the PDF.
    auto_manager = inv.get("auto_manager", "")
    manager_name = auto_manager or user.get("name", str(uid))
    sheet_ok = append_to_sheets(manager_name, inv)
    sheet_msg = "📊 Записано в Google Sheets ✓" if sheet_ok else "⚠️ Google Sheets недоступний"
    if auto_manager and auto_manager != user.get("name", ""):
        sheet_msg += f"\n👤 Лист менеджера: *{auto_manager}* (з таблиці)"

    ctx.user_data.pop("pending_invoice", None)
    ctx.user_data.pop("stock_selected", None)

    total_profit = 0
    total_revenue = 0.0
    for item in inv.get("items", []):
        lu = lookup_article(item["article"])
        cost_eur = lu.get("cost_eur", 0)
        duty = lu.get("duty", 0.04)
        rate = item.get("rate", 52.0)
        cost_uah = cost_eur * (1 + duty) * rate * item["qty"]
        revenue = item["price_uah"] * item["qty"]
        total_revenue += revenue
        profit = (item["price_uah"] - cost_eur * (1 + duty) * rate) * item["qty"] if item.get("is_stock") else revenue - cost_uah
        total_profit += profit

    # ── Перевірка оплат у момент завантаження рахунку ─────────────────────────
    # Тут сума рахунку (виторг) відома напряму з PDF, тому її НЕ треба шукати в «ВСІ».
    # Зберігаємо її як джерело правди й одразу дивимось: скільки оплат, на яку суму,
    # чи закрито, і чи треба коригування 50/50.
    inv_num = _inv_num(inv.get("invoice_num", ""))
    pay_block = ""
    if inv_num and total_revenue > 0:
        INVOICE_TOTALS[inv_num] = round(total_revenue, 2)
        _save_invoice_totals()
        _ensure_payments_loaded()
        tr = get_payments(inv_num)
        if tr:
            paid = total_paid(inv_num)
            n = len(tr)
            closed = paid >= total_revenue * 0.995
            if closed:
                wr = weighted_rate(inv_num)
                fr = first_payment_rate(inv_num)
                if wr > 0 and fr > 0 and abs(wr - fr) > 1e-6:
                    korig = round(total_revenue * (wr / fr - 1), 2)
                    total_profit += korig
                    pay_block = (
                        f"\n\n💳 Оплат: {n}, разом {paid:,.0f} грн — *закрито*\n"
                        f"Курс 1-ї опл.: {fr:.2f} → зважений: {wr:.2f}\n"
                        f"⚖️ Коригування 50/50: *{korig:+,.0f} грн* (у прибутку)"
                    ).replace(",", " ")
                else:
                    pay_block = (
                        f"\n\n💳 Оплат: {n}, разом {paid:,.0f} грн — *закрито* "
                        f"(один курс, коригування не потрібне)"
                    ).replace(",", " ")
            else:
                pay_block = (
                    f"\n\n💳 Оплат: {n}, разом {paid:,.0f} з {total_revenue:,.0f} грн — *часткова*\n"
                    f"⚠️ У ЗП піде лише після повної оплати"
                ).replace(",", " ")
        else:
            pay_block = "\n\n💳 Оплат по цьому рахунку ще немає"

    stock_count = len(selected)
    stock_msg = f"🔴 Складських: {stock_count}" if stock_count else "🟢 Складських немає"

    await query.edit_message_text(
        f"✅ *Рахунок збережено!*\n\n"
        f"💰 Прибуток: *{total_profit:,.0f} грн*\n"
        f"{stock_msg}\n"
        f"📁 Рахунків цього місяця: {len(user.get('invoices', []))}\n"
        f"{sheet_msg}"
        f"{pay_block}\n\n"
        f"Надішли наступний PDF або /report для Excel.",
        parse_mode="Markdown"
    )

async def cmd_report(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введи витрати на доставку (грн) або /0 якщо немає:")
    return WAIT_DELIVERY

def read_sheet_rows_for_manager(manager_name: str):
    """Read all data rows from manager's Google Sheet for current month."""
    try:
        sh = get_gsheet()
        if not sh:
            return None, "no_connection"
        try:
            ws = sh.worksheet(manager_name)
        except gspread.WorksheetNotFound:
            return [], "no_sheet"
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return [], "empty"
        # Columns: A..X (0..23). Data rows have article in col E (idx 4)
        # Skip header (row 1), separator rows (empty), and invoice-title rows (no article)
        rows = []
        for r in all_vals[1:]:
            if len(r) < 24:
                continue
            article = r[4].strip()
            if not article:
                continue  # separator or title row
            rows.append(r)
        return rows, "ok"
    except Exception as e:
        logger.error(f"read_sheet error: {e}")
        return None, str(e)

def _dt(s):
    try:
        return datetime.strptime(s, "%d.%m.%Y")
    except:
        return datetime.min

def _num_cell(v):
    try:
        return float(str(v).replace("\xa0","").replace(" ","").replace(",","."))
    except:
        return 0.0

def _group_rows_by_invoice(all_vals):
    """Group a manager/ВСІ sheet's rows into invoices.
    Returns list of {'inv': num, 'rows': [item rows]} (header/separator rows excluded)."""
    groups = []
    cur = None
    for r in all_vals[1:]:
        if len(r) < 5:
            continue
        cnum = r[2].strip() if len(r) > 2 else ""
        art = r[4].strip() if len(r) > 4 else ""
        if art:
            if cur is None:
                cur = {"inv": "", "rows": []}
                groups.append(cur)
            cur["rows"].append(r)
        elif cnum:
            m = re.search(r"[№#No]+\s*(\d+)", cnum)
            cur = {"inv": m.group(1) if m else "", "rows": []}
            groups.append(cur)
        # fully-empty separator row → ignore
    return [g for g in groups if g["rows"]]

def read_invoices_for_manager(manager_name: str):
    """Read manager's sheet grouped by invoice. Returns (groups, status)."""
    try:
        sh = get_gsheet()
        if not sh:
            return None, "no_connection"
        try:
            ws = sh.worksheet(manager_name)
        except gspread.WorksheetNotFound:
            return [], "no_sheet"
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return [], "empty"
        return _group_rows_by_invoice(all_vals), "ok"
    except Exception as e:
        logger.error(f"read_invoices error: {e}")
        return None, str(e)

def invoice_revenue_map_all() -> dict:
    """Map invoice_num -> total revenue (виторг).

    Primary source = totals captured from the PDF at upload time
    (INVOICE_TOTALS). The 'ВСІ' sheet is only a best-effort fallback for
    invoices processed before this was introduced.
    """
    out = {}
    try:
        sh = get_gsheet()
        if sh:
            try:
                ws = sh.worksheet("ВСІ")
                for g in _group_rows_by_invoice(ws.get_all_values()):
                    if not g["inv"]:
                        continue
                    rev = sum(_num_cell(r[12]) for r in g["rows"] if len(r) > 12)
                    out[g["inv"]] = round(out.get(g["inv"], 0) + rev, 2)
            except gspread.WorksheetNotFound:
                pass
    except Exception as e:
        logger.error(f"revenue map error: {e}")
    # Totals captured from PDF win over anything read from the sheet.
    for inv, tot in INVOICE_TOTALS.items():
        try:
            if tot and float(tot) > 0:
                out[str(inv)] = round(float(tot), 2)
        except:
            pass
    return out

def _load_payments_from_sheet():
    """Durable store = 'Оплати' tab. Load it into PAYMENTS (sheet wins)."""
    global _payments_loaded_from_sheet
    _payments_loaded_from_sheet = True
    try:
        sh = get_gsheet()
        if not sh:
            return
        try:
            ws = sh.worksheet("Оплати")
        except gspread.WorksheetNotFound:
            return
        vals = ws.get_all_values()
        loaded = {}
        for r in vals[1:]:
            if len(r) < 2:
                continue
            inv = re.sub(r"\D", "", str(r[0]))
            date = str(r[1]).strip() if len(r) > 1 else ""
            if not inv or not date:
                continue
            amount = _num_cell(r[2]) if len(r) > 2 else 0.0
            rate = _num_cell(r[3]) if len(r) > 3 else 0.0
            loaded.setdefault(inv, []).append(
                {"date": date, "amount": round(amount, 2), "rate": round(rate, 4)})
        if loaded:
            for inv, lst in loaded.items():
                PAYMENTS[inv] = lst
            _save_payments()
    except Exception as e:
        logger.error(f"load payments from sheet error: {e}")

def _ensure_payments_loaded():
    if not _payments_loaded_from_sheet:
        _load_payments_from_sheet()

def _write_payments_sheet(revenue_map: dict) -> bool:
    """Rebuild the 'Оплати' tab from PAYMENTS with accumulation + status."""
    sh = get_gsheet()
    if not sh:
        return False
    try:
        try:
            ws = sh.worksheet("Оплати")
            ws.clear()
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title="Оплати", rows=2000, cols=8)
        header = ["Рахунок", "Дата оплати", "Сума оплати", "Курс",
                  "Накопичено", "Сума рахунку", "Статус"]
        rows = [header]
        for inv in sorted(PAYMENTS.keys(), key=lambda x: (len(x), x)):
            tr = sorted(PAYMENTS[inv], key=lambda x: _dt(x.get("date", "")))
            total = revenue_map.get(inv, 0)
            acc = 0.0
            for t in tr:
                acc += float(t.get("amount", 0) or 0)
                if total > 0:
                    status = "закрито" if acc >= total * 0.995 else "часткова"
                elif acc > 0:
                    status = "невідома сума рах."
                else:
                    status = ""
                rows.append([inv, t.get("date", ""),
                             round(float(t.get("amount", 0) or 0), 2),
                             round(float(t.get("rate", 0) or 0), 4),
                             round(acc, 2),
                             round(total, 2) if total else "",
                             status])
        ws.update(f"A1:G{len(rows)}", rows, value_input_option="USER_ENTERED")
        ws.format("A1:G1", {
            "backgroundColor": {"red": 0.17, "green": 0.18, "blue": 0.24},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "horizontalAlignment": "CENTER"
        })
        return True
    except Exception as e:
        logger.error(f"write payments sheet error: {e}")
        return False

async def handle_delivery(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace("/","")
    try:
        delivery = float(text.replace(",",".").replace(" ",""))
    except:
        delivery = 0

    uid = update.effective_user.id
    user = load_user(uid)
    name = user.get("name", str(uid))
    month = datetime.now().strftime("%m.%Y")

    _ensure_payments_loaded()

    # Read manager's sheet grouped by invoice
    groups, status = read_invoices_for_manager(name)
    if status == "no_connection":
        await update.message.reply_text("⚠️ Google Sheets недоступний. Спробуй пізніше.")
        return ConversationHandler.END
    if status in ("no_sheet", "empty") or not groups:
        await update.message.reply_text("📭 Немає рахунків цього місяця в таблиці.")
        return ConversationHandler.END

    def num(v):
        return _num_cell(v)

    counted_rows = []      # item rows of invoices that count toward salary
    total_revenue = 0.0
    total_profit = 0.0
    open_invoices = []     # [(inv, paid, total)]
    closed_breakdown = []  # [(inv, revenue, wrate, profit)]

    for g in groups:
        inv = g["inv"]
        rows = g["rows"]
        revenue = sum(num(r[12]) for r in rows)                     # M Виторг
        profit_old = sum(num(r[13]) + num(r[14]) for r in rows)     # N+O (база)

        tranches = get_payments(inv) if inv else []
        if not tranches:
            # No payment data → behave as before
            counted_rows.extend(rows)
            total_revenue += revenue
            total_profit += profit_old
            continue

        paid = round(sum(float(t.get("amount", 0) or 0) for t in tranches), 2)
        amounts_known = any(float(t.get("amount", 0) or 0) > 0 for t in tranches)
        is_closed = (not amounts_known) or (revenue <= 0) or (paid >= revenue * 0.995)

        if not is_closed:
            open_invoices.append((inv, paid, round(revenue, 2)))
            continue

        # Currency-difference коригування: база за курсом 1-ї оплати,
        # пізніше оплачена частка дорожчає за своїм курсом.
        first_rate = first_payment_rate(inv)
        wr = weighted_rate(inv)
        korig = 0.0
        if first_rate > 0 and wr > 0:
            korig = round(revenue * (wr / first_rate - 1), 2)
        profit = round(profit_old + korig, 2)

        counted_rows.extend(rows)
        total_revenue += round(revenue + korig, 2)
        total_profit += profit
        closed_breakdown.append((inv, round(revenue, 2), first_rate, wr, korig, profit))

    net = total_profit - delivery
    salary = max(0, net) * SALARY_PCT / 100

    path = build_excel_from_rows(name, counted_rows, month, delivery, total_revenue,
                                 total_profit, net, salary, closed_breakdown, open_invoices)

    open_msg = ""
    if open_invoices:
        open_msg = f"\n⏳ Відкритих (не в ЗП): *{len(open_invoices)}*"

    await update.message.reply_document(
        document=open(path,"rb"),
        filename=f"ЗП_{name}_{month}.xlsx",
        caption=(
            f"📊 *Звіт за {month}*\n\n"
            f"Позицій (зараховано): {len(counted_rows)}\n"
            f"Виторг: *{total_revenue:,.0f} грн*\n"
            f"Прибуток: *{total_profit:,.0f} грн*\n"
            f"Доставка: *{delivery:,.0f} грн*\n"
            f"Чистий: *{net:,.0f} грн*"
            f"{open_msg}\n"
            f"━━━━━━━━━━━━\n"
            f"💰 *ЗП ({SALARY_PCT}%): {salary:,.0f} грн*"
        ),
        parse_mode="Markdown"
    )
    return ConversationHandler.END

def build_excel_from_rows(manager_name, rows, month, delivery, total_revenue, total_profit, net, salary,
                          closed_breakdown=None, open_invoices=None):
    """Build Excel report from Google Sheet rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month
    thin = Side(style='thin', color='BDC3C7')
    brd = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells('A1:N1')
    ws['A1'] = f'РОЗРАХУНОК ЗП — {manager_name} — {month}'
    ws['A1'].font = Font(name='Arial',bold=True,size=13,color='FFFFFF')
    ws['A1'].fill = PatternFill('solid',fgColor='1F2D3D')
    ws['A1'].alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 26

    headers = ['Клієнт','Рахунок','Дата','Артикул','Кть','Закуп EUR','Курс',
               'Ціна UAH','Виторг','Прибуток S','Надбавка T','Склад?','Постачальник']
    ws.row_dimensions[2].height = 30
    for c,h in enumerate(headers,1):
        cell = ws.cell(2,c,h)
        cell.font = Font(name='Arial',bold=True,size=9,color='FFFFFF')
        cell.fill = PatternFill('solid',fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = brd
    for i,w in enumerate([18,20,11,22,5,10,8,12,13,13,13,8,24],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def num(v):
        try: return float(str(v).replace("\xa0","").replace(" ","").replace(",","."))
        except: return 0.0

    row = 3
    for r in rows:
        is_stock = (len(r) > 15 and r[15] == "так")
        supplier = r[24] if len(r) > 24 else ""
        vals = [r[1], r[2], r[3], r[4], num(r[5]), num(r[6]), num(r[8]),
                num(r[11]), num(r[12]), num(r[13]), num(r[14]),
                "так" if is_stock else "", supplier]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row,c,v)
            cell.font = Font(name='Arial',size=9)
            cell.border = brd
            if c in [5,6,7,8,9,10,11]:
                cell.number_format = '#,##0.00'
            cell.fill = PatternFill('solid',fgColor='FDECEA' if is_stock else ('F8F9FA' if row%2==0 else 'FFFFFF'))
        row += 1

    TOT = row
    ws.merge_cells(f'A{TOT}:H{TOT}')
    ws.cell(TOT,1,'ПІДСУМОК').font = Font(name='Arial',bold=True,size=10,color='FFFFFF')
    ws.cell(TOT,1).fill = PatternFill('solid',fgColor='1A5276')
    ws.cell(TOT,1).alignment = Alignment(horizontal='right')
    for c in [9,10,11]:
        cl = get_column_letter(c)
        cell = ws.cell(TOT,c,f'=SUM({cl}3:{cl}{TOT-1})')
        cell.font = Font(name='Arial',bold=True,size=10,color='FFFFFF')
        cell.fill = PatternFill('solid',fgColor='1A5276')
        cell.number_format = '#,##0.00'

    SR = TOT+2
    labels = [('Прибуток (S+T) грн:', total_profit),
              ('Доставка грн:', delivery),
              ('Чистий прибуток грн:', net),
              (f'ЗП ({SALARY_PCT}%) грн:', salary)]
    for i,(lbl,val) in enumerate(labels):
        r2=SR+i
        ws.merge_cells(f'A{r2}:C{r2}')
        ws.cell(r2,1,lbl).font=Font(name='Arial',size=10,bold=True)
        ws.cell(r2,1).alignment=Alignment(horizontal='right')
        cv=ws.cell(r2,4,round(val,2))
        cv.number_format='#,##0.00'
        if i==3:
            cv.fill=PatternFill('solid',fgColor='D5F5E3')
            cv.font=Font(name='Arial',size=14,color='1E8449',bold=True)

    nxt = SR + len(labels) + 1

    # Closed invoices with currency-difference коригування (paid across >1 rate)
    cb = [b for b in (closed_breakdown or []) if b[4]]  # only those with a non-zero коригування
    if cb:
        ws.merge_cells(f'A{nxt}:F{nxt}')
        ws.cell(nxt,1,'КОРИГУВАННЯ (курсова різниця за оплатами)').font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        ws.cell(nxt,1).fill=PatternFill('solid',fgColor='1A5276')
        for c,h in enumerate(['Рахунок','Виторг','Курс 1-ї опл.','Курс зваж.','Коригування','Прибуток'],1):
            cell=ws.cell(nxt+1,c,h)
            cell.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
            cell.fill=PatternFill('solid',fgColor='2C3E50'); cell.border=brd
        rr=nxt+2
        for inv,rev,fr,wr,kor,pr in cb:
            for c,v in enumerate([f'рах.{inv}', round(rev,2), round(fr,4), round(wr,4), round(kor,2), round(pr,2)],1):
                cell=ws.cell(rr,c,v); cell.font=Font(name='Arial',size=9); cell.border=brd
                if c in [2,3,4,5,6]: cell.number_format='#,##0.00'
            rr+=1
        nxt = rr + 1

    # Open (partially paid) invoices — NOT in salary
    oi = open_invoices or []
    if oi:
        ws.merge_cells(f'A{nxt}:E{nxt}')
        ws.cell(nxt,1,'ВІДКРИТІ РАХУНКИ (не повністю оплачені — НЕ в ЗП)').font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        ws.cell(nxt,1).fill=PatternFill('solid',fgColor='922B21')
        for c,h in enumerate(['Рахунок','Оплачено','Сума рахунку','Залишок'],1):
            cell=ws.cell(nxt+1,c,h)
            cell.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
            cell.fill=PatternFill('solid',fgColor='C0392B'); cell.border=brd
        rr=nxt+2
        for inv,paid,tot in oi:
            for c,v in enumerate([f'рах.{inv}', round(paid,2), round(tot,2), round(tot-paid,2)],1):
                cell=ws.cell(rr,c,v); cell.font=Font(name='Arial',size=9); cell.border=brd
                if c in [2,3,4]: cell.number_format='#,##0.00'
                cell.fill=PatternFill('solid',fgColor='FDEDEC')
            rr+=1

    path = str(DATA_DIR/f"salary_{manager_name}_{month}.xlsx")
    wb.save(path)
    return path

async def cmd_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton("✅ Так",callback_data="clear_yes"),
                 InlineKeyboardButton("❌ Ні",callback_data="clear_no")]]
    await update.message.reply_text("⚠️ Видалити всі рахунки місяця?",
                                     reply_markup=InlineKeyboardMarkup(keyboard))

async def callback_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "clear_yes":
        user = load_user(query.from_user.id)
        user["invoices"] = []
        save_user(query.from_user.id, user)
        await query.edit_message_text("✅ Очищено!")
    else:
        await query.edit_message_text("Скасовано.")

async def cmd_admin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return
    lines = [f"📋 *Менеджери цього місяця:*\n"]
    for f in DATA_DIR.glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            name = data.get("name", f.stem)
            invoices = data.get("invoices",[])
            profit = 0
            for inv in invoices:
                import re as re_a
                m_a = re_a.search(r"[№#No]+\s*(\d+)", inv.get("invoice_num",""))
                inv_num_a = m_a.group(1) if m_a else ""
                for item in inv.get("items",[]):
                    lu = lookup_article(item["article"], inv_num_a, item.get("is_stock", False))
                    cost_eur = lu.get("cost_eur",0)
                    duty = lu.get("duty",0.04)
                    rate = item.get("rate",52.0)
                    cost_uah = cost_eur*(1+duty)*rate*item["qty"]
                    revenue = item["price_uah"]*item["qty"]
                    p = (item["price_uah"]-cost_eur*(1+duty)*rate)*item["qty"] if item.get("is_stock") else revenue-cost_uah
                    profit += p
            salary = max(0,profit)*SALARY_PCT/100
            lines.append(f"👤 *{name}*: {len(invoices)} рахунків\n   Прибуток: {profit:,.0f} грн | ЗП: {salary:,.0f} грн\n")
        except: pass
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введи нове ім'я:")
    return WAIT_NAME

async def cmd_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin only: update lookup from Excel file"""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return
    await update.message.reply_text(
        "📂 Надішли Excel файл таблиці замовлень (.xlsx)\n"
        "Аркуші мають називатись: *E-Trade Automation* та *Китай*",
        parse_mode="Markdown"
    )
    return WAIT_EXCEL

async def cmd_rates(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin only: update currency rates from Excel file"""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return
    await update.message.reply_text(
        "📂 Надішли Excel файл з курсами валют (.xlsx)\n"
        "Формат: колонка *Дата* і *Курс покупки* (або аналогічна структура)",
        parse_mode="Markdown"
    )
    return WAIT_RATES

async def handle_rates_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Process uploaded rates Excel"""
    if update.effective_user.id not in ADMIN_IDS:
        return ConversationHandler.END

    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Потрібен .xlsx файл")
        return WAIT_RATES

    msg = await update.message.reply_text("⏳ Оновлюю курси...")
    try:
        import pandas as pd

        file = await ctx.bot.get_file(doc.file_id)
        xlsx_path = str(DATA_DIR / f"rates_{doc.file_id}.xlsx")
        await file.download_to_drive(xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        df = pd.read_excel(xl, sheet_name=0, header=1)
        df.columns = [str(c).strip() for c in df.columns]
        import os as os_mod
        os_mod.remove(xlsx_path)

        # Find date and buy columns
        date_col = next((c for c in df.columns if 'дат' in c.lower() or 'date' in c.lower()), df.columns[0])
        # Accept both buy/sell column names - just take the rate column (2nd column)
        buy_col = next(
            (c for c in df.columns if 'покуп' in c.lower() or 'buy' in c.lower() or 'курс' in c.lower()),
            df.columns[1]
        )

        rates_new = {}
        skipped_junk = 0
        for _, row in df.iterrows():
            try:
                d = row[date_col]
                if pd.isna(d): continue
                # skip junk/placeholder dates (e.g. 01.01.1900 from a malformed cell)
                if hasattr(d, "year") and d.year < 2000:
                    skipped_junk += 1
                    continue
                date_str = d.strftime("%d.%m.%Y") if hasattr(d, "strftime") else str(d)[:10]
                if date_str.endswith("1900"):
                    skipped_junk += 1
                    continue
                buy = float(re.sub(r"[^\d.,\-]", "", str(row[buy_col])).replace(",", "."))
                if buy > 0:
                    rates_new[date_str] = buy
            except: pass

        with open("rates.json", "w", encoding="utf-8") as f:
            json.dump(rates_new, f, ensure_ascii=False, indent=2)

        RATES.clear()
        RATES.update(rates_new)

        await msg.edit_text(
            f"✅ *Курси оновлено!*\n\n"
            f"📅 Дат: *{len(rates_new)}*\n"
            f"🗑 Пропущено сміття: {skipped_junk}\n"
            f"📆 Від {min(rates_new.keys())} до {max(rates_new.keys())}",
            parse_mode="Markdown"
        )
    except Exception as e:
        await msg.edit_text(f"❌ Помилка: {e}")

    return ConversationHandler.END

async def handle_excel_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Process uploaded Excel and rebuild all 3 lookup files"""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return ConversationHandler.END

    doc = update.message.document
    if not doc.file_name.endswith('.xlsx'):
        await update.message.reply_text("❌ Потрібен .xlsx файл")
        return WAIT_EXCEL

    msg = await update.message.reply_text("⏳ Оновлюю довідники (lookup + invoice + stock)...")

    try:
        import pandas as pd
        import re as re_up
        from datetime import datetime as dt_up

        file = await ctx.bot.get_file(doc.file_id)
        xlsx_path = str(DATA_DIR / f"orders_{doc.file_id}.xlsx")
        await file.download_to_drive(xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        sheets = xl.sheet_names

        import os as os_mod

        # Sheet-specific invoice column names
        invoice_col_names = {
            'E-Trade Automation': '№ Рахунку',
            'Китай': 'Рахунок Україна',
        }

        new_lookup = {}
        new_invoice_lookup = {}
        new_stock_lookup = {}
        new_invoice_meta = {}

        for sheet in sheets:
            try:
                df = pd.read_excel(xl, sheet_name=sheet, header=0)
            except Exception as e:
                logger.warning(f"Sheet {sheet} error: {e}")
                continue

            cols = list(df.columns)
            def find_col(keywords):
                for kw in keywords:
                    for i, c in enumerate(cols):
                        if kw.lower() in str(c).lower(): return i
                return -1

            art_i = find_col(['артикул'])
            price_i = find_col(['ціна за одиницю', 'price'])
            uktved_i = find_col(['код товару', 'уктзед'])
            duty_i = find_col(['мито'])
            weight_i = find_col(['нетто за 1', 'вага'])
            brand_i = find_col(['виробник', 'brand'])
            supplier_i = find_col(['постачальник', 'поставщик', 'supplier'])
            manager_i = find_col(['менеджер', 'manager'])
            paydate_i = find_col(['дата оплати клієнт', 'дата оплати', 'дата оплаты'])
            date_i = find_col(['дата підтвердження'])
            inv_col_name = invoice_col_names.get(sheet, '№ Рахунку')
            invoice_i = find_col([inv_col_name])

            if art_i < 0: continue

            for _, row in df.iterrows():
                art = str(row.iloc[art_i]).strip()
                if not art or art in ('nan', 'Артикул', ''): continue

                try:
                    raw_price = str(row.iloc[price_i]) if price_i >= 0 else "0"
                    cleaned = re_up.sub(r"[^\d.,\-]", "", raw_price).replace(",", ".")
                    price = float(cleaned) if cleaned else 0
                except: price = 0

                try:
                    duty = float(str(row.iloc[duty_i]).replace(",",".").replace(" ","").replace("%","")) if duty_i >= 0 else 0.04
                    if duty > 1: duty = duty / 100
                except: duty = 0.04
                try: weight = float(str(row.iloc[weight_i]).replace(",",".").replace(" ","")) if weight_i >= 0 else 0
                except: weight = 0
                try: brand = str(row.iloc[brand_i]).strip() if brand_i >= 0 else ''
                except: brand = ''
                brand = '' if brand == 'nan' else brand
                try: supplier = str(row.iloc[supplier_i]) if supplier_i >= 0 else ''
                except: supplier = ''
                # strip openpyxl control-char escapes (_x000D_, _x0002_ ...) and collapse spaces
                supplier = re_up.sub(r'_x[0-9A-Fa-f]{4}_', ' ', supplier)
                supplier = re_up.sub(r'\s+', ' ', supplier).strip()
                supplier = '' if supplier in ('nan', 'NaN', 'None', '') else shorten_company(supplier)

                # Manager (skip "Склад" placeholder)
                try: manager = str(row.iloc[manager_i]).strip() if manager_i >= 0 else ''
                except: manager = ''
                if manager in ('nan', 'NaN', 'None', '') or 'склад' in manager.lower():
                    manager = ''

                # Client payment date (AA) — actual date client paid
                pay_date = ''
                pay_obj = None
                if paydate_i >= 0:
                    try:
                        pd_v = row.iloc[paydate_i]
                        if pd.notna(pd_v) and hasattr(pd_v, 'strftime'):
                            pay_date = pd_v.strftime('%d.%m.%Y')
                            pay_obj = pd_v
                    except: pass

                date_str = ''
                date_obj = None
                if date_i >= 0:
                    try:
                        d = row.iloc[date_i]
                        if pd.notna(d) and hasattr(d, 'strftime'):
                            date_str = d.strftime('%d.%m.%Y')
                            date_obj = d
                    except: pass

                inv_raw = str(row.iloc[invoice_i]).strip() if invoice_i >= 0 else ''
                is_storage = 'склад' in inv_raw.lower()
                inv_num = ''
                if not is_storage:
                    m = re_up.search(r'(?:№|No|#)\s*(\d+)', inv_raw, re_up.IGNORECASE)
                    if m: inv_num = m.group(1)

                record = {
                    'uktved': str(row.iloc[uktved_i]).strip() if uktved_i >= 0 and uktved_i < len(row) else '',
                    'duty': duty,
                    'cost_eur': price,
                    'weight': weight,
                    'brand': brand,
                    'source': sheet,
                    'confirm_date': date_str,
                    'invoice_num': inv_num,
                    'is_storage': is_storage,
                    'supplier': supplier,
                    'manager': manager,
                    'pay_date': pay_date,
                }

                # Stock lookup
                if is_storage and price > 0:
                    existing = new_stock_lookup.get(art)
                    if not existing:
                        new_stock_lookup[art] = record
                    else:
                        try:
                            ex_date = dt_up.strptime(existing['confirm_date'], '%d.%m.%Y') if existing.get('confirm_date') else dt_up.min
                            new_date = date_obj.to_pydatetime() if hasattr(date_obj, 'to_pydatetime') else (date_obj if date_obj else dt_up.min)
                            if new_date > ex_date:
                                new_stock_lookup[art] = record
                        except:
                            new_stock_lookup[art] = record

                # Invoice lookup
                if inv_num:
                    key = f"{inv_num}:{art}"
                    if key not in new_invoice_lookup or (price > 0 and new_invoice_lookup[key].get('cost_eur', 0) == 0):
                        new_invoice_lookup[key] = record

                # Invoice-level meta (manager / pay_date / supplier) keyed by invoice number.
                # On number collision across years keep the record with the LATEST date.
                if inv_num:
                    try:
                        cand = pay_obj or date_obj
                        cand_dt = cand.to_pydatetime() if hasattr(cand, 'to_pydatetime') else cand
                    except:
                        cand_dt = None
                    cur = new_invoice_meta.get(inv_num)
                    if cur is None:
                        new_invoice_meta[inv_num] = {
                            'manager': manager, 'pay_date': pay_date,
                            'supplier': supplier, '_dt': cand_dt,
                        }
                    else:
                        newer = False
                        try:
                            newer = bool(cand_dt) and (cur.get('_dt') is None or cand_dt > cur['_dt'])
                        except:
                            newer = False
                        if newer:
                            new_invoice_meta[inv_num] = {
                                'manager': manager or cur.get('manager', ''),
                                'pay_date': pay_date or cur.get('pay_date', ''),
                                'supplier': supplier or cur.get('supplier', ''),
                                '_dt': cand_dt,
                            }
                        else:
                            if not cur.get('manager') and manager: cur['manager'] = manager
                            if not cur.get('pay_date') and pay_date: cur['pay_date'] = pay_date
                            if not cur.get('supplier') and supplier: cur['supplier'] = supplier

                # General lookup
                existing = new_lookup.get(art, {})
                if price > 0 or not existing:
                    record_copy = dict(record)
                    if price == 0 and existing.get('cost_eur', 0) > 0:
                        record_copy['cost_eur'] = existing['cost_eur']
                    new_lookup[art] = record_copy

        os_mod.remove(xlsx_path)

        # Strip internal sort key before saving meta
        clean_meta = {
            k: {kk: vv for kk, vv in v.items() if kk != '_dt'}
            for k, v in new_invoice_meta.items()
        }

        # Save all files
        import json as json_mod
        with open('lookup.json', 'w', encoding='utf-8') as f:
            json_mod.dump(new_lookup, f, ensure_ascii=False, indent=2)
        with open('invoice_lookup.json', 'w', encoding='utf-8') as f:
            json_mod.dump(new_invoice_lookup, f, ensure_ascii=False, indent=2)
        with open('stock_lookup.json', 'w', encoding='utf-8') as f:
            json_mod.dump(new_stock_lookup, f, ensure_ascii=False, indent=2)
        with open('invoice_meta.json', 'w', encoding='utf-8') as f:
            json_mod.dump(clean_meta, f, ensure_ascii=False, indent=2)

        # Update in-memory
        LOOKUP.clear(); LOOKUP.update(new_lookup)
        INVOICE_LOOKUP.clear(); INVOICE_LOOKUP.update(new_invoice_lookup)
        STOCK_LOOKUP.clear(); STOCK_LOOKUP.update(new_stock_lookup)
        INVOICE_META.clear(); INVOICE_META.update(clean_meta)

        meta_with_date = sum(1 for v in clean_meta.values() if v.get('pay_date'))
        meta_with_mgr = sum(1 for v in clean_meta.values() if v.get('manager'))
        await msg.edit_text(
            f"✅ *Довідники оновлено!*\n\n"
            f"📦 Артикулів: *{len(new_lookup)}*\n"
            f"📄 По рахунках: *{len(new_invoice_lookup)}*\n"
            f"📦 Склад: *{len(new_stock_lookup)}*\n"
            f"👤 Рахунків з менеджером: *{meta_with_mgr}*\n"
            f"💸 Рахунків з датою оплати: *{meta_with_date}*\n"
            f"📋 Аркушів: {len(sheets)} ({', '.join(sheets[:3])})",
            parse_mode="Markdown"
        )

    except Exception as e:
        logger.error(f"Excel update error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        await msg.edit_text(f"❌ Помилка: {e}")

    return ConversationHandler.END


# ── Payments (actual client payment dates from Teams) ─────────────────────────
async def cmd_oplata(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin: load actual payment dates. Accepts pasted Teams text OR an .xlsx."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return ConversationHandler.END
    await update.message.reply_text(
        "💸 Надішли оплати — *встав текст* як з Teams, або кинь *Excel файл*.\n\n"
        "Формат (як у Teams):\n"
        "`Оплати за 29/05/26 «...»`\n"
        "`1. КЛІЄНТ рах.834 – 64 531,20`\n"
        "`2. КЛІЄНТ рах.835 – 33 799,20`\n\n"
        "Сума важлива: бот рахує часткові оплати. Рахунок «закрито», коли оплачено ≥ суми рахунку; "
        "курс для собівартості — зважений за датами/сумами оплат.",
        parse_mode="Markdown"
    )
    return WAIT_PAYMENTS

async def _finish_payments(update, tranches: list):
    if not tranches:
        await update.message.reply_text(
            "❌ Не знайшов жодного рядка. Перевір формат: заголовок «Оплати за ДД/ММ/РР» "
            "і рядки з `рах.НОМЕР – СУМА`.",
            parse_mode="Markdown"
        )
        return
    _ensure_payments_loaded()
    # Compute rate per tranche date (archive → Minfin → NBU)
    rate_cache = {}
    for t in tranches:
        d = t["date"]
        if d not in rate_cache:
            rate_cache[d] = await get_payment_rate(d) or 0.0
        t["rate"] = rate_cache[d]

    added = _merge_payment_tranches(tranches)

    # Persist to the durable 'Оплати' tab (with accumulation + status vs invoice total)
    revenue_map = invoice_revenue_map_all()
    sheet_ok = _write_payments_sheet(revenue_map)

    invs = sorted({t["inv"] for t in tranches}, key=lambda x: (len(x), x))
    lines = [f"✅ *Оплати збережено:* +{added} трансакцій по {len(invs)} рах."]
    for inv in invs[:8]:
        paid = total_paid(inv)
        tot = revenue_map.get(inv, 0)
        if tot > 0:
            st = "закрито" if paid >= tot * 0.995 else "часткова"
            lines.append(f"• рах.{inv}: оплачено {paid:,.0f} з {tot:,.0f} — {st}".replace(",", " "))
        else:
            lines.append(f"• рах.{inv}: оплачено {paid:,.0f} (сума рах. ще невідома)".replace(",", " "))
    if len(invs) > 8:
        lines.append(f"…та ще {len(invs) - 8} рах.")
    lines.append("\n📊 Вкладка *Оплати* оновлена ✓" if sheet_ok else "\n⚠️ Google Sheets недоступний (збережено локально)")
    lines.append("Курс для закритих рахунків — зважений за сумами оплат. Частково оплачені в ЗП не йдуть.")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def handle_payments_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return ConversationHandler.END
    text = update.message.text or ""
    tranches = _parse_payment_text(text)
    await _finish_payments(update, tranches)
    return ConversationHandler.END

async def handle_payments_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return ConversationHandler.END
    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Потрібен .xlsx файл або встав текст.")
        return WAIT_PAYMENTS
    msg = await update.message.reply_text("⏳ Читаю оплати...")
    try:
        file = await ctx.bot.get_file(doc.file_id)
        xlsx_path = str(DATA_DIR / f"pay_{doc.file_id}.xlsx")
        await file.download_to_drive(xlsx_path)

        # Read every cell, join into text lines, reuse the same text parser
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        text_lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    for ln in str(cell).splitlines():
                        text_lines.append(ln)
        os.remove(xlsx_path)

        tranches = _parse_payment_text("\n".join(text_lines))
        await msg.delete()
        await _finish_payments(update, tranches)
    except Exception as e:
        logger.error(f"oplata excel error: {e}")
        await msg.edit_text(f"❌ Помилка: {e}")
    return ConversationHandler.END


# ── Manual invoice totals (Сума рахунку) for old invoices ────────────────────
async def cmd_suma(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin: set invoice totals (виторг) by hand for invoices not in the bot."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return ConversationHandler.END
    await update.message.reply_text(
        "🧾 Введи суми рахунків (виторг), по рядку на рахунок:\n"
        "`65 = 146268`\n`66 = 29886`\n\n"
        "Можна й через пробіл: `65 146268`.\n"
        "Це потрібно для старих рахунків, яких немає в боті — щоб він зрозумів, "
        "коли рахунок «закрито» і чи треба коригування 50/50.",
        parse_mode="Markdown"
    )
    return WAIT_SUMA

def _parse_amount(s: str) -> float:
    s = str(s).replace(" ", "").replace("\xa0", "")
    if s.count(",") and s.count("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

async def handle_suma_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return ConversationHandler.END
    text = update.message.text or ""
    added = {}
    for line in text.splitlines():
        # "рах.65 = 146268", "65: 146268", "65 - 146268" or "65 146268"
        m = re.match(r"\s*(?:рах\.?\s*№?\s*)?(\d+)\s*[=:\-]\s*([\d\s.,]+)", line, re.IGNORECASE) \
            or re.match(r"\s*(?:рах\.?\s*№?\s*)?(\d+)\s+([\d][\d\s.,]{2,})", line, re.IGNORECASE)
        if not m:
            continue
        inv = m.group(1)
        val = _parse_amount(m.group(2))
        if val > 0:
            INVOICE_TOTALS[str(inv)] = round(val, 2)
            added[inv] = round(val, 2)
    if not added:
        await update.message.reply_text(
            "❌ Не розпізнав жодного рядка. Формат: `65 = 146268`",
            parse_mode="Markdown"
        )
        return WAIT_SUMA
    _save_invoice_totals()
    _ensure_payments_loaded()
    sheet_ok = _write_payments_sheet(invoice_revenue_map_all())
    lines = [f"✅ *Збережено сум рахунків: {len(added)}*"]
    for inv, v in list(added.items())[:10]:
        paid = total_paid(inv)
        if paid > 0:
            st = "закрито" if paid >= v * 0.995 else "часткова"
            lines.append(f"• рах.{inv}: сума {v:,.0f}, оплачено {paid:,.0f} → {st}".replace(",", " "))
        else:
            lines.append(f"• рах.{inv}: сума {v:,.0f} (оплат ще немає)".replace(",", " "))
    if len(added) > 10:
        lines.append(f"…та ще {len(added) - 10}")
    lines.append("\n📊 Вкладка *Оплати* оновлена ✓" if sheet_ok else "\n⚠️ Google Sheets недоступний (збережено локально)")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    return ConversationHandler.END

# ── Recompute all stored payment rates from the archive/NBU ──────────────────
async def cmd_perekurs(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin: re-resolve the курс of every stored tranche (unfreezes old 52,72)."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Немає доступу.")
        return
    _ensure_payments_loaded()
    msg = await update.message.reply_text("⏳ Перераховую курси по всіх збережених оплатах...")
    rate_cache = {}
    changed = 0
    for inv, lst in PAYMENTS.items():
        for t in lst:
            d = t.get("date", "")
            if not d:
                continue
            if d not in rate_cache:
                rate_cache[d] = await get_payment_rate(d) or 0.0
            nr = round(float(rate_cache[d]), 4)
            if nr > 0 and round(float(t.get("rate", 0) or 0), 4) != nr:
                t["rate"] = nr
                changed += 1
    _save_payments()
    sheet_ok = _write_payments_sheet(invoice_revenue_map_all())
    await msg.edit_text(
        f"✅ *Курси перераховано.*\n"
        f"🔁 Оновлено траншів: *{changed}*\n"
        + ("📊 Вкладка *Оплати* оновлена ✓" if sheet_ok else "⚠️ Google Sheets недоступний"),
        parse_mode="Markdown"
    )


async def cmd_sheet(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if SHEET_ID:
        await update.message.reply_text(
            f"📊 Google таблиця:\nhttps://docs.google.com/spreadsheets/d/{SHEET_ID}"
        )

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(BOT_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Document.PDF, handle_pdf),
            CommandHandler("report", cmd_report),
            CommandHandler("name", cmd_name),
            CommandHandler("update", cmd_update),
            CommandHandler("rates", cmd_rates),
            CommandHandler("oplata", cmd_oplata),
            CommandHandler("suma", cmd_suma),
        ],
        states={
            WAIT_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, set_name)],
            WAIT_DATE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date)],
            WAIT_STOCK:    [
                CallbackQueryHandler(callback_stock, pattern="^stk_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_stock),
            ],
            WAIT_DELIVERY: [MessageHandler(filters.TEXT, handle_delivery),
                            CommandHandler("0", handle_delivery)],
            WAIT_RATES: [MessageHandler(filters.Document.MimeType(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ), handle_rates_update)],
            WAIT_EXCEL: [MessageHandler(filters.Document.MimeType(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ), handle_excel_update)],
            WAIT_PAYMENTS: [
                MessageHandler(filters.Document.MimeType(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ), handle_payments_excel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payments_text),
            ],
            WAIT_SUMA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_suma_text)],
        },
        fallbacks=[CommandHandler("start", start)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(CommandHandler("clear", cmd_clear))
    app.add_handler(CommandHandler("admin", cmd_admin))
    app.add_handler(CommandHandler("sheet", cmd_sheet))
    app.add_handler(CommandHandler("perekurs", cmd_perekurs))
    app.add_handler(CallbackQueryHandler(callback_clear, pattern="^clear_"))
    logger.info("Bot started")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
